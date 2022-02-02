import smtplib
import email,sys,json
import pathlib
from email.mime.text import MIMEText            # 负责构造文本
from email.mime.image import MIMEImage          # 负责构造图片
from email.mime.multipart import MIMEMultipart  # 负责将多个对象集合起来
from email.header import Header

from openpyxl import Workbook

from datetime import datetime, timedelta
from tempfile import NamedTemporaryFile
import os

print(str(pathlib.Path.cwd().parent.parent.parent))
sys.path.append(str(pathlib.Path.cwd().parent.parent.parent))      # 解决命令找不到包路径
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "biddingapi.settings")  # manage.py文件中有同样的环境配置
import django
django.setup()  # 启动django项目
from django.db.models import F, Q
sys.path.append(str(pathlib.Path.cwd().parent.parent))      # 解决命令找不到包路径
from users.models import User
from bid.models import Bidding
from django.utils.http import urlquote





def get_file():
    user_data = User.objects.filter(state=True).all().values()
    print(user_data,len(user_data))
    get_file_data = []
    for i in user_data:
        user_name = i['username']
        key_word = i['key_word']
        url_id = i['url_id']
        time_day = i['time_day']
        email = i['email']


        if not key_word:
            key_word = '第三方、满意度、调查、统计、调研、检查、研究、咨询、巡查、普查、考核、测评、评估、绩效、创建、摸底、核查、入户、监测、社会救助、城市管理'
        print(url_id,type(url_id))
     
        if not url_id:
            value_u = ['0']
        elif type(url_id) != type([]) and '[' not in url_id:
            value_u = [str(url_id)]
            #print(value_u,'elif********************')
        else:
            value_u = eval(url_id)

        if not time_day:
            time_day = '1'

        #print(value_u,'******************************')

        s_time = datetime.now()

        value_time = [(s_time - timedelta(days=time_day)).strftime('%Y-%m-%d %H:%M:%S'),s_time.strftime('%Y-%m-%d %H:%M:%S')]

        if key_word:
            key_word = '|'.join(key_word.replace("，", ',').replace("、", ',').split(','))
        else:
            key_word = '.*?'
        print(user_name, key_word, value_u, value_time, email)
        if '0' in value_u or value_u is None or value_u == []:
            if value_time:
                data_list = Bidding.objects.filter(
                    b_title__regex=r'{}'.format(key_word),
                    b_release__range=[value_time[0], value_time[1]], ).all().values(
                    'collect_time', 'b_title', 'b_url', 'b_release',
                    'b_money', 'customer_name', 'customer_phone',
                    'b_time', 'collect__web_name'
                )
            else:

                data_list = Bidding.objects.filter(
                    # b_title__regex=r'{}'.format(key_word),
                ).all().values(
                    'collect_time', 'b_title', 'b_url', 'b_release',
                    'b_money', 'customer_name', 'customer_phone',
                    'b_time', 'collect__web_name'
                )
        else:
            if value_time:
                data_list = Bidding.objects.filter(
                    b_title__regex=r'{}'.format(key_word),
                    collect_id__in=value_u, b_release__range=(
                        value_time[0],
                        value_time[1])
                ).all().values(
                    'collect_time', 'b_title', 'b_url', 'b_release',
                    'b_money', 'customer_name', 'customer_phone',
                    'b_time', 'collect__web_name'
                )
            else:

                data_list = Bidding.objects.filter(
                    b_title__regex=r'{}'.format(key_word),
                    collect_id__in=value_u
                ).all().values(
                    'collect_time', 'b_title', 'b_url', 'b_release',
                    'b_money', 'customer_name', 'customer_phone',
                    'b_time', 'collect__web_name'
                )

        # print(data_list,len(data_list))
        d_count = len(data_list)
        workbook = Workbook()
        # wb = Workbook()
        ws = workbook.active
        # ws = wb.create_sheet("Mysheet")  # insert at the end (default)
        # ws = wb['Mysheet']
        ws['A1'] = '序号'
        ws['B1'] = '采集日期'
        ws['C1'] = '招标标题'
        ws['D1'] = '网页链接地址'
        ws['E1'] = '发标日期'
        ws['F1'] = '招标金额'
        ws['G1'] = '客户名称'
        ws['H1'] = '客户联系方式'
        ws['I1'] = '获取招标文件时间'
        ws['J1'] = '来源网站名'
        s1_time = datetime.now()
        for index_i, i in enumerate(data_list, start=1):
            i['collect_time'] = i['collect_time'].strftime('%Y-%m-%d %H:%M:%S')
            i['b_time'] = i['b_time'].strftime('%Y-%m-%d %H:%M:%S')
            i['b_release'] = i['b_release'].strftime('%Y-%m-%d %H:%M:%S')
            ws_data = list(i.values())

            ws_data.insert(0, index_i)
            # print(ws_data)
            # print(i)
            ws.append(ws_data)
        print(datetime.now() - s1_time)
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp)
            tmp.seek(0)
            stream = tmp.read()



        get_file_data.append([stream,user_name,email,d_count])
    return get_file_data

def send_email(f_name,subject_content, body_content,mail_receivers):
    """
    f_name: 文件的字节流
    subject_content: 邮件主题
    body_content: 邮件文本内容
    :return:
    """
    # 构建MIMEMultipart对象代表邮件本身，可以往里面添加文本、图片、附件等
    mm = MIMEMultipart('related')

    # 邮件主题
    subject_content = subject_content

    # 设置发送者,注意严格遵守格式,里面邮箱为发件人邮箱
    mm["From"] = "Yanilo<{}>".format(mail_sender)

    # 设置接受者,注意严格遵守格式,里面邮箱为接受者邮箱
    mm["To"] = "Destiny<{}>".format(mail_receivers)
    # 设置邮件主题

    mm["Subject"] = Header(subject_content, 'utf-8')

    # 邮件正文内容
    body_content = body_content

    # 构造文本,参数1：正文内容，参数2：文本格式，参数3：编码方式
    message_text = MIMEText(body_content, "plain", "utf-8")

    # 向MIMEMultipart对象中添加文本对象
    mm.attach(message_text)

    # 构造附件
    atta = MIMEText(f_name, 'base64', 'utf-8')


    # file_name_data = urlquote("招标信息")+datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    # print(file_name_data,type(file_name_data))
    # 设置附件信息

    atta.add_header("Content-Disposition", "attachment", filename=("utf-8", "", '招标信息{}.xlsx'.format(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))))


    # 添加附件到邮件信息当中去
    mm.attach(atta)

    # 创建SMTP对象
    stp = smtplib.SMTP_SSL("smtp.qq.com", 465)

    # set_debuglevel(1)可以打印出和SMTP服务器交互的所有信息
    # stp.set_debuglevel(1)

    # 登录邮箱，传递参数1：邮箱地址，参数2：邮箱授权码
    stp.login(mail_sender, 'psldyaxwwltqbgbd')

    # 发送邮件，传递参数1：发件人邮箱地址，参数2：收件人邮箱地址，参数3：把邮件内容格式改为str
    stp.sendmail(mail_sender, mail_receivers, mm.as_string())

    print("邮件发送成功！")

    # 关闭SMTP对象
    stp.quit()
    return 'ok'


# SMTP服务器,这里使用qq邮箱
mail_host = "smtp.qq.com"

# 发件人邮箱
mail_sender = "39017764@qq.com"

# 邮箱授权码,注意这里不是邮箱密码,如何获取邮箱授权码
mail_license = "复制的授权码"

# 收件人邮箱，可以为多个收件人
# mail_receivers = "xiao3952@foxmail.com"

# mail_receivers = ["xiao3901@foxmail.com",'xx@xx.com'] # 多个收件人

# 构建MIMEMultipart对象代表邮件本身，可以往里面添加文本、图片、附件等
mm = MIMEMultipart('related')





try:
    with open('pid.json','r+',encoding='utf-8') as fp:
        j_pid = json.loads(fp.read())
except:
    j_pid = {'pid':[os.getpid()]}
    with open('pid.json','w+',encoding='utf-8') as fp:
        fp.write(json.dumps(j_pid))
print(j_pid,type(j_pid))
with open('pid.json','w',encoding='utf-8') as fp1:
    j_pid['pid'].append(os.getpid())
    fp1.write(json.dumps(j_pid))

def cron_send():
    d_list = get_file()
    for u_data in d_list:

        mail_receivers = u_data[2]
        f_name = u_data[0]  # 本地文件名
        subject_content = """招标信息邮件"""  # 邮件主题
        body_content = """Hello there,{} ！
            招标信息更新啦！ 这次一共更新了{}条。
    
        Don't have a good day,have a great day！
        """.format(u_data[1], u_data[-1])  # 邮件文本内容
        send_email(f_name, subject_content, body_content,mail_receivers)
    return 'ok'

# cron_send()

import time
from apscheduler.schedulers.blocking import BlockingScheduler


def my_job():
    cron_send()
    print(os.getpid())
    print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())))


#sched = BlockingScheduler(timezone='Asia/Shanghai')
#sched.add_job(my_job, 'cron',second = '*/5')
#sched.add_job(my_job, 'cron',day_of_week='*',hour='17,8',minute=34,second=10)
#sched.start()


#cron_send()
