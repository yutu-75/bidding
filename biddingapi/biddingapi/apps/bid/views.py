import json
from django.http import StreamingHttpResponse  # 流式传输一个响应给浏览器
from django.shortcuts import render, HttpResponse, redirect
# from spider_bidding.yancao import get_comment
# Create your views here.
from django.views import View
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from pathlib import Path
from openpyxl import Workbook
from tempfile import NamedTemporaryFile
from datetime import datetime
from django.utils.http import urlquote
from django.db.models import F, Q


from . import models


# from bidding.spider_bidding.www_ccgp_hubei_gov_cn import main

# def excel_get():
#
#     return stream


def index(request):
    pass



class LogoutView(APIView):
    def get(self, request):
        # 删除所有当前请求相关的session
        request.session.delete()
        return redirect("/bidding/login/")


class LoginView(APIView):
    def get(self, request):
        response_data = response_data = [{'label': "默认", 'value': '默认'}]

        return render(request, 'login.html', {'url_ip': BASE_URL_IP})

    def post(self, request):

        u_data = request.data
        print(u_data)
        uname = u_data['uname']

        password = u_data['password']

        u_bool = models.User.objects.filter(user_name=uname, password=password).first()
        print(bool(u_bool))
        if u_bool:
            # 设置session
            request.session["user"] = uname
            print(uname)
            # return render(request, 'index1.html', {'url_ip': BASE_URL_IP,'uname':uname})
            return Response(data={'err': 'ok'}, status=200)

        else:
            # error
            return Response(data={'err': '账号或密码错误！'}, status=400)

        # return HttpResponse(json.dumps('ok!      <a href=/bidding/login/>点击这里重新登录！</a>',ensure_ascii=False))


class RegisterView(APIView):
    def get(self, request):
        response_data = response_data = [{'label': "默认", 'value': '默认'}]

        return render(request, 'register.html', {'url_ip': BASE_URL_IP})

    def post(self, request):
        u_data = request.data
        print(u_data)

        uname = u_data['uname']
        email = u_data['email']
        password = u_data['password']

        judge = [i for i in [uname, email, password] if 0 < len(i) < 30]

        if len(judge) < 3:
            return Response(data={'err': '注册失败,有参数没有填或参数过长(不能超过30)！'}, status=400)

        uname_bool = models.User.objects.filter(user_name=uname).first()
        if uname_bool:
            return Response(data={'err': '注册失败,用户名已被使用！ '}, status=400)

        u_obj = models.User.objects.create(
            user_name=uname,
            email=email,
            password=password,
        )
        print(u_obj.user_name)  # 可以基于这个对象来取这个新添加的记录对象的属性值

        return Response(data={'err': 'ok'}, status=200)




class IndexView(APIView):
    permission_classes = [IsAuthenticated, ]
    def get(self, request):


        return Response('ok')

    def post(self, request):
        # print('wwwwwwwwwwwwwwwwwwww')

        # return Response('ok!!!!')
        req = request.data
        value_u = req.get('value_u', None)
        print(value_u,type(value_u))

        value_time = req.get('value_time', None)
        f_number = req.get('f_number',1)  # 当前页码
        p_number = req.get('p_number',20)  # 当前页码数据条数
        key_word = req.get('key_word', None)
        customer_name = req.get('customer_name', None)
        download = req.get('download', None)
        x = (f_number - 1) * p_number
        y = f_number * p_number
        if download == 'ture':



            x = None
            y = None


        if key_word:

            key_word = '|'.join(key_word.replace("，", ',').replace("、", ',').split(','))
        else:
            key_word = '.*?'

        if not customer_name:
            customer_name = '.*?'

        # 查看各个参数类型
        # aaa = [value_u, value_time, f_number, p_number, key_word]
        # print('----------')
        # for ia in aaa:
        #     print(ia)
        #     if ia:
        #         print(type(ia))
        # print('----------')

        if value_u is None or '0' in value_u  or value_u == []:
            if value_time:
                count_t = models.Bidding.objects.filter(
                    b_title__regex=r'{}'.format(key_word),
                    customer_name__regex=r'{}'.format(customer_name),
                    b_release__range=[value_time[0], value_time[1]]).count()
                data_list = models.Bidding.objects.filter(
                    b_title__regex=r'{}'.format(key_word),
                    customer_name__regex=r'{}'.format(customer_name),
                    b_release__range=[value_time[0], value_time[1]], ).all().values(
                    'collect_time', 'b_title', 'b_url', 'b_release',
                    'b_money', 'customer_name', 'customer_phone',
                    'b_time', 'collect__web_name'
                )[x:y]

            else:
                data_list = models.Bidding.objects.filter(
                    customer_name__regex=r'{}'.format(customer_name),
                    b_title__regex=r'{}'.format(key_word),
                ).all().values(
                    'collect_time', 'b_title', 'b_url', 'b_release',
                    'b_money', 'customer_name', 'customer_phone',
                    'b_time', 'collect__web_name'
                )[x:y]
                count_t = models.Bidding.objects.filter(
                    customer_name__regex=r'{}'.format(customer_name),
                    b_title__regex=r'{}'.format(key_word),
                ).count()


        # elif value_u is None or not value_u.isdigit():
        #     return Response('参数错误！')
        else:

            if value_time:
                count_t = models.Bidding.objects.filter(
                    collect_id__in=value_u,
                    customer_name__regex=r'{}'.format(customer_name),
                    b_release__range=(value_time[0], value_time[1]),
                    b_title__regex=r'{}'.format(key_word)
                ).count()
                data_list = models.Bidding.objects.filter(
                    collect_id__in=value_u,
                    customer_name__regex=r'{}'.format(customer_name),
                    b_title__regex=r'{}'.format(key_word),
                    b_release__range=(
                        value_time[0],
                        value_time[1]),
                ).all().values(
                    'collect_time', 'b_title', 'b_url', 'b_release',
                    'b_money', 'customer_name', 'customer_phone',
                    'b_time', 'collect__web_name'
                )[x:y]
            else:
                count_t = models.Bidding.objects.filter(
                    collect_id__in=value_u,
                    # b_title__contains=key_word,
                    b_title__regex=r'{}'.format(key_word),
                    customer_name__regex=r'{}'.format(customer_name),
                ).count()
                data_list = models.Bidding.objects.filter(
                    collect_id__in=value_u,
                    b_title__regex=r'{}'.format(key_word),
                    customer_name__regex=r'{}'.format(customer_name),

                ).all().values(
                    'collect_time', 'b_title', 'b_url', 'b_release',
                    'b_money', 'customer_name', 'customer_phone',
                    'b_time', 'collect__web_name'
                )[x:y]

        # data_list_1 = []
        # for i in data_list:
        #     i['collect_id'] = models.Collect.objects.filter(id=i['collect_id']).first().web_name
        #     data_list_1.append(i)

        # print(data_list)



        if download == 'ture':
            return self.e_download(data_list)
        else:


            response_data = {
                'data_list': data_list,
                'count_t': count_t,

            }

            return Response(response_data)

    @staticmethod
    def e_download(data_list):

        s_time = datetime.now()
        if not data_list:
            return Response(data={'err': '没有查询到数据！'}, status=400)
        print(datetime.now() - s_time)
        workbook = Workbook()
        # wb = Workbook()
        ws = workbook.active
        # ws = wb.create_sheet("Mysheet")  # insert at the end (default)
        # ws = wb['Mysheet']
        ws['A1'] = '序号'
        ws['B1'] = '采集日期'
        ws['C1'] = '招标标题'
        ws['D1'] = '网页链接地址'
        ws['E1'] = '发标日期'
        ws['F1'] = '招标金额'
        ws['G1'] = '客户名称'
        ws['H1'] = '客户联系方式'
        ws['I1'] = '获取招标文件时间'
        ws['J1'] = '来源网站名'
        s1_time = datetime.now()
        for index_i, i in enumerate(data_list, start=1):
            i['collect_time'] = i['collect_time'].strftime('%Y-%m-%d %H:%M:%S')
            i['b_time'] = i['b_time'].strftime('%Y-%m-%d %H:%M:%S')
            i['b_release'] = i['b_release'].strftime('%Y-%m-%d %H:%M:%S')
            ws_data = list(i.values())

            ws_data.insert(0, index_i)
            # print(ws_data)
            # print(i)
            ws.append(ws_data)
        print(datetime.now() - s1_time)
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(content=stream)
        response['Content-Type'] = 'application/octet-stream'  # 设置头信息，告诉浏览器这是个文件
        response[
            'Content-Disposition'] = f'attachment; filename={urlquote("招标信息")}-{datetime.now().strftime("%Y%m%d%H%M")}.xlsx'
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        print(datetime.now() - s_time)


        print(response['Content-Disposition'])
        return response





class WordView(APIView):
    def post(self, request):
        word_data = '第三方、满意度、调查、统计、调研、检查、研究、咨询、巡查、普查、考核、测评、评估、绩效、创建、摸底、核查、入户、监测、社会救助、城市管理'
        print(request.data)
        word_list = []
        for i, v in enumerate(word_data.split('、')):
            word_list.append({'id': i, 'name': v})
        return HttpResponse(json.dumps(word_list))





class BiddingView(APIView):

    def get(self, request):
        print('qwq')
        print(request)
        # main()
        return HttpResponse(json.dumps(['ok'], ensure_ascii=False))

    def post(self, request):
        print(request.data)
        return HttpResponse(json.dumps(['ok'], ensure_ascii=False))


class InputDataView(APIView):
    def get(self, request):
        pass
        return render(request, 'index.html')

    def post(self, request):
        pass

# class BiddingView(APIView):
#
#
#     def get(self,request):
#         print('qwq')
#         print(request)
#         return HttpResponse(json.dumps(['ok'],ensure_ascii=False))
#
#     def post(self,request):
#         print(request.POST)
#         print(request.data)
#         r_data = request.data
#         url = r_data['url']
#         word = r_data['word']
#         d_time = r_data['d_time']
#         email = r_data['email']
#         response_data, response_word = get_comment(url[-1])
#         print(response_data)
#         print(response_word)
#         response_data = {
#             'str_word': response_word,
#             'comment': response_data
#         }
#         # HttpResponse('ok')
#         return HttpResponse(json.dumps(response_data, ensure_ascii=False))
